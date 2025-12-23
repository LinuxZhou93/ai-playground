# Psyche-X™ 数据导出模块
# 支持 Excel, PDF, CSV 格式

from typing import List, Dict
import json
from datetime import datetime

class DataExporter:
    """企业级数据导出工具"""
    
    @staticmethod
    def export_to_csv(data: List[Dict], filename: str = None) -> str:
        """导出为 CSV 格式"""
        if not data:
            return ""
        
        # 获取所有字段
        headers = list(data[0].keys())
        
        # 生成 CSV 内容
        csv_lines = [",".join(headers)]
        
        for row in data:
            values = [str(row.get(h, "")) for h in headers]
            csv_lines.append(",".join(values))
        
        csv_content = "\n".join(csv_lines)
        
        if filename:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(csv_content)
        
        return csv_content
    
    @staticmethod
    def export_users_to_excel(users: List[Dict]) -> bytes:
        """导出用户列表为 Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "用户列表"
            
            # 标题行
            headers = ["ID", "用户名", "邮箱", "年龄", "年级", "测评次数", "最后测评", "注册时间"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 数据行
            for row_idx, user in enumerate(users, 2):
                ws.cell(row=row_idx, column=1, value=user.get('id'))
                ws.cell(row=row_idx, column=2, value=user.get('username'))
                ws.cell(row=row_idx, column=3, value=user.get('email'))
                ws.cell(row=row_idx, column=4, value=user.get('age'))
                ws.cell(row=row_idx, column=5, value=user.get('grade'))
                ws.cell(row=row_idx, column=6, value=user.get('assessment_count'))
                ws.cell(row=row_idx, column=7, value=user.get('last_assessment'))
                ws.cell(row=row_idx, column=8, value=user.get('created_at'))
            
            # 调整列宽
            for col in range(1, 9):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            # 保存到内存
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except ImportError:
            # 如果没有 openpyxl，返回 CSV 格式
            return DataExporter.export_to_csv(users).encode('utf-8')
    
    @staticmethod
    def export_report_to_pdf(report_data: Dict, user_info: Dict) -> bytes:
        """导出个人报告为 PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # 标题
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#0066CC'),
                spaceAfter=30,
                alignment=1  # 居中
            )
            
            story.append(Paragraph("Psyche-X™ 认知评估报告", title_style))
            story.append(Spacer(1, 20))
            
            # 用户信息
            user_data = [
                ["用户名", user_info.get('username', 'N/A')],
                ["邮箱", user_info.get('email', 'N/A')],
                ["测评日期", datetime.now().strftime("%Y-%m-%d")],
            ]
            
            user_table = Table(user_data, colWidths=[150, 300])
            user_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            
            story.append(user_table)
            story.append(Spacer(1, 30))
            
            # 认知能力得分
            if 'scores' in report_data:
                scores = report_data['scores']
                score_data = [
                    ["维度", "得分", "等级"],
                    ["流体智力 (Gf)", f"{scores.get('Gf', 0):.1f}", report_data.get('grades', {}).get('Gf', 'N/A')],
                    ["工作记忆 (Gwm)", f"{scores.get('Gwm', 0):.1f}", report_data.get('grades', {}).get('Gwm', 'N/A')],
                    ["执行功能 (Att)", f"{scores.get('Att', 0):.1f}", report_data.get('grades', {}).get('Att', 'N/A')],
                    ["元认知 (Meta)", f"{scores.get('Meta', 0):.1f}", report_data.get('grades', {}).get('Meta', 'N/A')],
                    ["心理韧性 (Res)", f"{scores.get('Res', 0):.1f}", report_data.get('grades', {}).get('Res', 'N/A')],
                ]
                
                score_table = Table(score_data, colWidths=[200, 150, 100])
                score_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(Paragraph("认知能力评估", styles['Heading2']))
                story.append(Spacer(1, 12))
                story.append(score_table)
            
            # 生成 PDF
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()
            
        except ImportError:
            # 如果没有 reportlab，返回 JSON 格式
            return json.dumps(report_data, ensure_ascii=False, indent=2).encode('utf-8')
    
    @staticmethod
    def export_batch_assessments(assessments: List[Dict], format: str = 'excel') -> bytes:
        """批量导出测评记录"""
        if format == 'csv':
            return DataExporter.export_to_csv(assessments).encode('utf-8')
        else:
            return DataExporter.export_users_to_excel(assessments)


# 使用示例
if __name__ == "__main__":
    # 测试数据
    test_users = [
        {"id": 1, "username": "student001", "email": "s1@example.com", "age": 15, "grade": "高一", "assessment_count": 12, "last_assessment": "2024-12-10", "created_at": "2024-11-01"},
        {"id": 2, "username": "student002", "email": "s2@example.com", "age": 16, "grade": "高二", "assessment_count": 8, "last_assessment": "2024-12-09", "created_at": "2024-11-05"},
    ]
    
    # 导出 CSV
    csv_content = DataExporter.export_to_csv(test_users)
    print("CSV Export:")
    print(csv_content)
    
    # 导出 Excel
    try:
        excel_data = DataExporter.export_users_to_excel(test_users)
        print(f"\nExcel Export: {len(excel_data)} bytes")
    except Exception as e:
        print(f"\nExcel Export failed: {e}")
